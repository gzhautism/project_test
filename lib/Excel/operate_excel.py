# -*- encoding: utf-8 -*-
# @File    : test.py
# @Contact : chengcheng@heading.loc
# @Modify Time : 2021/12/28 17:48
# @Author : chengchengy
# @Version : 1.0
# @Description: 操作Excel的基类

import openpyxl


class Excel:

    def __init__(self, file_path=None, sheet_id=None, sheet_name=None):
        if file_path:
            self.file_path = file_path
            self.sheet_id = sheet_id
        else:
            self.file_path = "operate_excel.elsx"

        if sheet_id:
            self.sheet_id = sheet_id
        else:
            self.sheet_id = 0

        if sheet_name:
            self.sheet_name = sheet_name
        else:
            self.sheet_name = "Sheet1"

        self.wb = openpyxl.load_workbook(self.file_path)

        self.sheet_table = self.get_sheet()

    def get_sheet(self):
        if self.sheet_name:
            sheet_table = self.wb[self.sheet_name]

        else:
            sheet_names = self.wb[self.sheet_name]
            sheet_table = self.wb[sheet_names[self.sheet_id]]

        return sheet_table

    @property
    def rows(self):
        i = self.sheet_table.max_row
        real_max_row = 0
        while i > 0:
            row_dict = {i.value for i in self.sheet_table[i]}
            if row_dict == {None}:
                i = i - 1
            else:
                real_max_row = i
                break

        return real_max_row

    @property
    def cols(self):
        i = self.sheet_table.max_column
        real_max_cols = 0
        while i > 0:
            cols_dict = {i for i in self.get_cell_value(self.rows, i)}
            if cols_dict == {None}:
                i = i - 1
            else:
                real_max_cols = i
                break

        return real_max_cols

    @property
    def title(self):
        return self.sheet_table.title

    def save_excel(self, file_name=None):
        if file_name is None:
            self.wb.save(self.file_path)
        else:
            self.wb.save(file_name)

    def create_sheet(self, title=None, index=None):
        if index is None:
            ws = self.wb.create_sheet(title=title)
        else:
            ws = self.wb.create_sheet(title=title, index=index)

        self.save_excel()

    def get_cell_value(self, row, col):
        return self.sheet_table.cell(row, col).value

    def insert_value_by_index(self, value, row, col):
        self.sheet_table.cell(row, col).value = value
        self.save_excel()

    def cell(self, row, col):
        return self.sheet_table.cell(row, col)
